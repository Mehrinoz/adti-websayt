import csv
import io
import json

from django import forms
from django.contrib import admin, messages
from django.db.models import Max
from django.http import HttpRequest, HttpResponseRedirect
from django.urls import path, reverse
from django.utils.safestring import mark_safe

from .models import Question, TestTuri, PracticeQuestion

# Word (DOCX) fayllari uchun importni yanada ishonchli qilish
try:
    import docx

    Document = docx.Document
except Exception:  # noqa: BLE001
    Document = None

try:  # Excel (xlsx) fayllari uchun
    from openpyxl import load_workbook
except Exception:  # noqa: BLE001
    load_workbook = None


class QuestionInline(admin.TabularInline):
    model = Question
    extra = 1
    readonly_fields = ("group_number",)


class QuestionUploadForm(forms.Form):
    """
    CSV yoki JSON faylni yuklash uchun forma.

    Kutilgan maydonlar:
      - question_text
      - choice_a
      - choice_b
      - choice_c
      - choice_d
      - correct_answer (A/B/C/D)
    """

    file = forms.FileField(label="Savollar fayli (CSV, JSON, DOCX yoki Excel)")
    chunk_size = forms.IntegerField(
        label="Har bo'limdagi savollar soni",
        min_value=1,
        initial=20,
        help_text="Masalan, 20 deb kiritsangiz, savollar 20 tadan bo'limlarga bo'linadi.",
    )


@admin.register(TestTuri)
class TestTuriAdmin(admin.ModelAdmin):
    list_display = ("name", "description")
    inlines = [QuestionInline]

    change_form_template = "admin/testapp/testturi/change_form.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:object_id>/upload-questions/",
                self.admin_site.admin_view(self.upload_questions_view),
                name="testapp_testturi_upload_questions",
            ),
        ]
        return custom_urls + urls

    def upload_questions_view(self, request: HttpRequest, object_id: int):
        category = self.get_object(request, object_id)
        if category is None:
            self.message_user(
                request,
                "Test turi topilmadi.",
                level=messages.ERROR,
            )
            return HttpResponseRedirect(
                reverse("admin:testapp_testturi_changelist")
            )

        if request.method == "POST":
            form = QuestionUploadForm(request.POST, request.FILES)
            if form.is_valid():
                upload_file = form.cleaned_data["file"]
                chunk_size = form.cleaned_data["chunk_size"]
                name = upload_file.name.lower()

                try:
                    if name.endswith(".csv"):
                        self._create_questions_from_csv(
                            upload_file, category, request, chunk_size
                        )
                    elif name.endswith(".json"):
                        self._create_questions_from_json(
                            upload_file, category, request, chunk_size
                        )
                    elif name.endswith(".docx"):
                        if Document is None:
                            self.message_user(
                                request,
                                "DOCX fayllarni o‘qish uchun avval 'python-docx' paketini o‘rnating.",
                                level=messages.ERROR,
                            )
                        else:
                            self._create_questions_from_docx(
                                upload_file, category, request, chunk_size
                            )
                    elif name.endswith(".xlsx"):
                        if load_workbook is None:
                            self.message_user(
                                request,
                                "Excel fayllarni o‘qish uchun avval 'openpyxl' paketini o‘rnating.",
                                level=messages.ERROR,
                            )
                        else:
                            self._create_questions_from_excel(
                                upload_file, category, request, chunk_size
                            )
                    else:
                        self.message_user(
                            request,
                            "Faqat CSV, JSON, DOCX yoki Excel (XLSX) fayl yuklash mumkin.",
                            level=messages.ERROR,
                        )
                        return HttpResponseRedirect(
                            reverse(
                                "admin:testapp_testturi_change",
                                args=[category.pk],
                            )
                        )
                except Exception as exc:  # noqa: BLE001
                    self.message_user(
                        request,
                        f"Faylni o‘qishda xatolik: {exc}",
                        level=messages.ERROR,
                    )

                return HttpResponseRedirect(
                    reverse(
                        "admin:testapp_testturi_change",
                        args=[category.pk],
                    )
                )
        else:
            form = QuestionUploadForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": category,
            "title": "Savollarni fayldan yuklash",
            "form": form,
        }
        from django.shortcuts import render

        return render(request, "admin/testapp/testturi/upload_questions.html", context)

    def _bulk_create_questions(
        self,
        rows: list[dict],
        category,
        request: HttpRequest,
        source: str,
        chunk_size: int,
    ):
        """
        Savollarni foydalanuvchi ko'rsatgan `chunk_size` bo'yicha bo'lib saqlash.
        """

        batch_size = chunk_size
        created = 0
        batch: list[Question] = []

        max_group = (
            Question.objects.filter(category=category).aggregate(
                max_group=Max("group_number")
            )["max_group"]
            or 0
        )
        current_group = max_group
        items_in_current_group = 0

        for row in rows:
            if items_in_current_group == 0:
                current_group += 1
            batch.append(
                Question(
                    category=category,
                    question_text=row["question_text"],
                    choice_a=row["choice_a"],
                    choice_b=row["choice_b"],
                    choice_c=row["choice_c"],
                    choice_d=row["choice_d"],
                    correct_answer=row["correct_answer"],
                    group_number=current_group,
                )
            )
            items_in_current_group += 1
            if items_in_current_group == chunk_size:
                items_in_current_group = 0

            if len(batch) == batch_size:
                Question.objects.bulk_create(batch)
                created += len(batch)
                batch = []

        if batch:
            Question.objects.bulk_create(batch)
            created += len(batch)

        self.message_user(
            request,
            f"{created} ta savol {source} fayldan muvaffaqiyatli qo‘shildi.",
            level=messages.SUCCESS,
        )

    def _create_questions_from_csv(
        self, upload_file, category, request: HttpRequest, chunk_size: int
    ):
        decoded = upload_file.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = []
        for row in reader:
            if not row.get("question_text"):
                continue
            rows.append(
                {
                    "question_text": row["question_text"],
                    "choice_a": row["choice_a"],
                    "choice_b": row["choice_b"],
                    "choice_c": row["choice_c"],
                    "choice_d": row["choice_d"],
                    "correct_answer": row["correct_answer"].upper(),
                }
            )

        self._bulk_create_questions(rows, category, request, "CSV", chunk_size)

    def _create_questions_from_json(
        self, upload_file, category, request: HttpRequest, chunk_size: int
    ):
        data = json.load(upload_file)
        if isinstance(data, dict):
            data = data.get("questions", [])

        rows = []
        for item in data:
            if not item.get("question_text"):
                continue
            rows.append(
                {
                    "question_text": item["question_text"],
                    "choice_a": item["choice_a"],
                    "choice_b": item["choice_b"],
                    "choice_c": item["choice_c"],
                    "choice_d": item["choice_d"],
                    "correct_answer": item["correct_answer"].upper(),
                }
            )

        self._bulk_create_questions(rows, category, request, "JSON", chunk_size)

    def _create_questions_from_docx(
        self, upload_file, category, request: HttpRequest, chunk_size: int
    ):
        """
        Word (DOCX) fayldan savollarni o‘qish.

        Kutilgan format (takrorlanadi):

        Savol matni
        A) variant A
        B) variant B
        C) variant C
        D) variant D
        To'g'ri: A

        Bo'sh qatordan keyin keyingi savol.
        """

        if Document is None:
            raise RuntimeError("python-docx o‘rnatilmagan.")

        document = Document(upload_file)
        lines = [p.text.strip() for p in document.paragraphs if p.text.strip()]

        def _clean_choice(line: str) -> str:
            # "A) matn" yoki "A. matn" bo'lsa, boshini olib tashlaymiz
            if len(line) > 2 and line[1] in [")", "."]:
                return line[2:].lstrip()
            return line

        rows = []
        i = 0
        n = len(lines)

        while i + 6 <= n:
            question_text = lines[i]
            a_line = lines[i + 1]
            b_line = lines[i + 2]
            c_line = lines[i + 3]
            d_line = lines[i + 4]
            correct_line = lines[i + 5]

            if not question_text or not correct_line:
                i += 1
                continue

            # Endi format: "Javob: A" yoki "Javob A"
            norm = correct_line.lower()
            if not norm.startswith("javob"):
                i += 1
                continue

            parts = correct_line.split(":", 1)
            if len(parts) == 2:
                candidate = parts[1].strip()
            else:
                # "Javob A" ko'rinishida bo'lsa
                candidate = correct_line[len("Javob") :].strip()

            correct = candidate.upper()[:1]
            if correct not in {"A", "B", "C", "D"}:
                i += 1
                continue

            rows.append(
                {
                    "question_text": question_text,
                    "choice_a": _clean_choice(a_line),
                    "choice_b": _clean_choice(b_line),
                    "choice_c": _clean_choice(c_line),
                    "choice_d": _clean_choice(d_line),
                    "correct_answer": correct,
                }
            )
            i += 6

        self._bulk_create_questions(
            rows, category, request, "DOCX (Word)", chunk_size
        )

    def _create_questions_from_excel(
        self, upload_file, category, request: HttpRequest, chunk_size: int
    ):
        """
        Excel (XLSX) fayldan savollarni o‘qish.

        Kutilgan ustunlar (birinchi qatorda sarlavha bo‘lishi mumkin):

        A: Savol
        B: To'g'ri javob
        C–E: Noto'g'ri javoblar
        """

        if load_workbook is None:
            raise RuntimeError("openpyxl o‘rnatilmagan.")

        workbook = load_workbook(upload_file, read_only=True, data_only=True)
        sheet = workbook.active

        def _cell_to_str(value) -> str:
            if value is None:
                return ""
            return str(value).strip()

        rows: list[dict] = []
        first = True
        for row in sheet.iter_rows(values_only=True):
            if first:
                # birinchi qator sarlavha deb hisoblaymiz
                first = False
                continue

            question_text = _cell_to_str(row[0] if len(row) > 0 else None)
            correct = _cell_to_str(row[1] if len(row) > 1 else None)
            wrong1 = _cell_to_str(row[2] if len(row) > 2 else None)
            wrong2 = _cell_to_str(row[3] if len(row) > 3 else None)
            wrong3 = _cell_to_str(row[4] if len(row) > 4 else None)

            if not question_text or not correct:
                continue

            # To'g'ri javobni A qilib, noto'g'ri javoblarni B, C, D ga joylaymiz
            choices = [correct, wrong1, wrong2, wrong3]
            # Bo'sh bo'lmaganlarini to'playmiz
            non_empty = [c for c in choices if c]
            if len(non_empty) < 2:  # kamida bitta noto'g'ri javob bo'lsin
                continue

            # Birinchi element to'g'ri javob bo'lib qoladi, qolganlari noto'g'ri
            choice_a = non_empty[0]
            other = non_empty[1:4]
            # Agar 3 tadan kam bo'lsa, qolganlarini bo'sh qoldiramiz
            while len(other) < 3:
                other.append("")

            rows.append(
                {
                    "question_text": question_text,
                    "choice_a": choice_a,
                    "choice_b": other[0],
                    "choice_c": other[1],
                    "choice_d": other[2],
                    "correct_answer": "A",
                }
            )

        self._bulk_create_questions(rows, category, request, "Excel (XLSX)", chunk_size)

    def render_change_form(self, request, context, *args, **kwargs):
        obj = context.get("original")
        if obj:
            upload_url = reverse(
                "admin:testapp_testturi_upload_questions", args=[obj.pk]
            )
            context["upload_link"] = mark_safe(
                f'<a class="button" href="{upload_url}">Savollarni fayldan yuklash</a>'
            )
        return super().render_change_form(request, context, *args, **kwargs)


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ("question_text", "category", "group_number", "correct_answer")
    list_filter = ("category", "group_number")
    search_fields = ("question_text",)


@admin.register(PracticeQuestion)
class PracticeQuestionAdmin(admin.ModelAdmin):
    list_display = ("question_text", "created_at", "updated_at")
    search_fields = ("question_text", "correct_answer")
    list_filter = ("created_at",)
    readonly_fields = ("created_at", "updated_at")
    fields = ("question_text", "correct_answer", "created_at", "updated_at")
